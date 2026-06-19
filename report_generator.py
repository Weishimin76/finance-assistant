# -*- coding: utf-8 -*-
"""
报表生成器模块 - report_generator.py
支持Excel输出和图表可视化
使用openpyxl生成Excel报表，使用plotly生成交互式图表
文件编码: UTF-8
"""

import os
import logging
from datetime import datetime
from typing import Optional, List, Dict, Any

import pandas as pd

# openpyxl 用于生成 Excel
try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers,
    )
    from openpyxl.utils import get_column_letter
    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False

# plotly 用于生成交互式图表
try:
    import plotly.graph_objects as go
    from plotly.subplots import make_subplots
    _PLOTLY_AVAILABLE = True
except ImportError:
    _PLOTLY_AVAILABLE = False

# ---------------------------------------------------------------------------
# 日志
# ---------------------------------------------------------------------------
logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# 样式常量
# ---------------------------------------------------------------------------
_HEADER_FONT = Font(name="Microsoft YaHei", bold=True, size=11, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_SUBHEADER_FONT = Font(name="Microsoft YaHei", bold=True, size=10, color="333333")
_SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
_DATA_FONT = Font(name="Microsoft YaHei", size=10)
_TITLE_FONT = Font(name="Microsoft YaHei", bold=True, size=14, color="1F4E79")
_SUBTITLE_FONT = Font(name="Microsoft YaHei", bold=True, size=12, color="2E75B6")
_NUMBER_FORMAT = '#,##0.00'
_PERCENT_FORMAT = '0.00%'
_THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
_CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
_RIGHT_ALIGN = Alignment(horizontal="right", vertical="center")

# 配色方案（用于图表）
_CHART_COLORS = [
    "#4472C4", "#ED7D31", "#A5A5A5", "#FFC000", "#5B9BD5",
    "#70AD47", "#264478", "#9B5700", "#7030A0", "#C55A11",
]


# ===========================================================================
# Excel 样式辅助
# ===========================================================================
def _apply_header_style(ws, row, max_col):
    """为表头行应用样式"""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER_ALIGN
        cell.border = _THIN_BORDER


def _apply_subheader_style(ws, row, max_col):
    """为子表头行应用样式"""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = _SUBHEADER_FONT
        cell.fill = _SUBHEADER_FILL
        cell.alignment = _CENTER_ALIGN
        cell.border = _THIN_BORDER


def _apply_data_style(ws, row, max_col, number_cols=None):
    """为数据行应用样式"""
    if number_cols is None:
        number_cols = []
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = _DATA_FONT
        cell.border = _THIN_BORDER
        if col in number_cols:
            cell.alignment = _RIGHT_ALIGN
            cell.number_format = _NUMBER_FORMAT
        else:
            cell.alignment = _LEFT_ALIGN


def _auto_width(ws, min_width=10, max_width=40):
    """自动调整列宽"""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                cell_len = len(str(cell.value))
                # 中文字符宽度约为英文的2倍
                cn_chars = sum(1 for c in str(cell.value) if '\u4e00' <= c <= '\u9fff')
                cell_len += cn_chars
                max_len = max(max_len, cell_len)
        adjusted_width = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted_width


def _safe_float(value, default=0.0):
    """安全转换为浮点数"""
    try:
        if pd.isna(value) or value is None:
            return default
        return float(value)
    except (ValueError, TypeError):
        return default


# ===========================================================================
# 1. 月度利润表 Excel 生成
# ===========================================================================
def generate_profit_report(df: pd.DataFrame, output_path: str) -> str:
    """
    生成月度利润Excel报表。

    Sheet1: 月度汇总（总收入、各平台收入、总成本、毛利润、毛利率、净利润、净利率）
    Sheet2: 平台明细
    Sheet3: SKU明细
    Sheet4: 公式说明

    Args:
        df: 包含订单数据的DataFrame，期望列：
            - order_date / date: 订单日期
            - platform: 平台名称
            - sku / product_sku: SKU编码
            - order_amount / amount: 订单金额
            - commission / fee: 佣金/手续费
            - shipping: 运费
            - refund: 退款金额
            - currency: 币种
        output_path: 输出Excel文件路径

    Returns:
        str: 生成的文件路径
    """
    if not _OPENPYXL_AVAILABLE:
        raise ImportError("需要安装 openpyxl 库: pip install openpyxl")

    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else ".", exist_ok=True)

    wb = Workbook()

    # ---- 预处理数据 ----
    df = df.copy()

    # 列名映射关系（不使用rename避免多列映射到同名的冲突）
    _COL_MAPPING = {
        "order_date": ["order_date", "date", "订单日期", "交易日期"],
        "platform": ["platform", "平台"],
        "sku": ["sku", "product_sku", "sku编码"],
        "order_amount": ["order_amount", "amount", "订单金额", "销售额"],
        "commission": ["commission", "fee", "佣金"],
        "shipping": ["shipping", "运费", "物流费"],
        "refund": ["refund", "退款", "退款金额"],
        "currency": ["currency", "币种"],
    }

    # 为每个标准列找到第一个匹配的原始列
    std_cols = {}
    for std_name, candidates in _COL_MAPPING.items():
        for c in candidates:
            if c in df.columns:
                std_cols[std_name] = c
                break

    # 创建标准化列（用新列名，不覆盖原始列）
    if "order_date" in std_cols:
        df["order_date"] = pd.to_datetime(df[std_cols["order_date"]], errors="coerce")
        df["month"] = df["order_date"].dt.to_period("M").astype(str)
    else:
        df["month"] = "未知月份"

    # 确保数值列
    for num_col in ["order_amount", "commission", "shipping", "refund"]:
        if num_col in std_cols:
            df[num_col] = pd.to_numeric(df[std_cols[num_col]], errors="coerce").fillna(0)
        else:
            df[num_col] = 0.0

    # 计算衍生列
    df["total_cost"] = df["commission"] + df["shipping"] + df["refund"]
    df["gross_profit"] = df["order_amount"] - df["commission"] - df["shipping"]
    df["net_profit"] = df["order_amount"] - df["total_cost"]

    # ==================================================================
    # Sheet1: 月度汇总
    # ==================================================================
    ws1 = wb.active
    ws1.title = "月度汇总"

    # 标题
    ws1.merge_cells("A1:H1")
    ws1["A1"] = "月度利润汇总报表"
    ws1["A1"].font = _TITLE_FONT
    ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws1.merge_cells("A2:H2")
    ws1["A2"] = f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws1["A2"].font = Font(name="Microsoft YaHei", size=9, color="666666")
    ws1["A2"].alignment = Alignment(horizontal="center")

    # 表头
    row = 4
    headers = ["月份", "总收入", "总佣金", "总运费", "总退款", "总成本",
               "毛利润", "毛利率", "净利润", "净利率", "订单数"]
    for col_idx, header in enumerate(headers, 1):
        ws1.cell(row=row, column=col_idx, value=header)
    _apply_header_style(ws1, row, len(headers))

    # 按月汇总
    monthly = df.groupby("month").agg(
        total_revenue=("order_amount", "sum"),
        total_commission=("commission", "sum"),
        total_shipping=("shipping", "sum"),
        total_refund=("refund", "sum"),
        order_count=("order_amount", "count"),
    ).reset_index()

    monthly["total_cost"] = monthly["total_commission"] + monthly["total_shipping"] + monthly["total_refund"]
    monthly["gross_profit"] = monthly["total_revenue"] - monthly["total_commission"] - monthly["total_shipping"]
    monthly["gross_margin"] = monthly.apply(
        lambda r: r["gross_profit"] / r["total_revenue"] if r["total_revenue"] > 0 else 0, axis=1
    )
    monthly["net_profit"] = monthly["total_revenue"] - monthly["total_cost"]
    monthly["net_margin"] = monthly.apply(
        lambda r: r["net_profit"] / r["total_revenue"] if r["total_revenue"] > 0 else 0, axis=1
    )

    number_cols = [2, 3, 4, 5, 6, 7, 10]
    percent_cols = [8, 11]

    for idx, m_row in monthly.iterrows():
        row += 1
        ws1.cell(row=row, column=1, value=str(m_row["month"]))
        ws1.cell(row=row, column=2, value=round(_safe_float(m_row["total_revenue"]), 2))
        ws1.cell(row=row, column=3, value=round(_safe_float(m_row["total_commission"]), 2))
        ws1.cell(row=row, column=4, value=round(_safe_float(m_row["total_shipping"]), 2))
        ws1.cell(row=row, column=5, value=round(_safe_float(m_row["total_refund"]), 2))
        ws1.cell(row=row, column=6, value=round(_safe_float(m_row["total_cost"]), 2))
        ws1.cell(row=row, column=7, value=round(_safe_float(m_row["gross_profit"]), 2))
        ws1.cell(row=row, column=8, value=_safe_float(m_row["gross_margin"]))
        ws1.cell(row=row, column=8).number_format = _PERCENT_FORMAT
        ws1.cell(row=row, column=9, value=round(_safe_float(m_row["net_profit"]), 2))
        ws1.cell(row=row, column=10, value=_safe_float(m_row["net_margin"]))
        ws1.cell(row=row, column=10).number_format = _PERCENT_FORMAT
        ws1.cell(row=row, column=11, value=int(m_row["order_count"]))
        _apply_data_style(ws1, row, len(headers), number_cols=number_cols)

    # 合计行
    row += 1
    ws1.cell(row=row, column=1, value="合计")
    ws1.cell(row=row, column=1).font = Font(name="Microsoft YaHei", bold=True, size=10)
    for col_idx in range(2, 12):
        if col_idx in percent_cols:
            # 计算加权平均
            if col_idx == 8:
                total_gp = monthly["gross_profit"].sum()
                total_rev = monthly["total_revenue"].sum()
                val = total_gp / total_rev if total_rev > 0 else 0
            else:
                total_np = monthly["net_profit"].sum()
                total_rev = monthly["total_revenue"].sum()
                val = total_np / total_rev if total_rev > 0 else 0
            ws1.cell(row=row, column=col_idx, value=val)
            ws1.cell(row=row, column=col_idx).number_format = _PERCENT_FORMAT
        elif col_idx == 11:
            ws1.cell(row=row, column=col_idx, value=int(monthly["order_count"].sum()))
        else:
            col_letter = get_column_letter(col_idx)
            formula = f"=SUM({col_letter}5:{col_letter}{row - 1})"
            ws1.cell(row=row, column=col_idx, value=formula)
            ws1.cell(row=row, column=col_idx).number_format = _NUMBER_FORMAT
        ws1.cell(row=row, column=col_idx).font = Font(name="Microsoft YaHei", bold=True, size=10)
        ws1.cell(row=row, column=col_idx).border = _THIN_BORDER

    _auto_width(ws1)

    # ==================================================================
    # Sheet2: 平台明细
    # ==================================================================
    ws2 = wb.create_sheet("平台明细")

    ws2.merge_cells("A1:I1")
    ws2["A1"] = "各平台收入明细"
    ws2["A1"].font = _TITLE_FONT
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")

    row = 3
    headers2 = ["平台", "订单数", "总收入", "总佣金", "总运费", "总退款",
                "毛利润", "毛利率", "净利润"]
    for col_idx, header in enumerate(headers2, 1):
        ws2.cell(row=row, column=col_idx, value=header)
    _apply_header_style(ws2, row, len(headers2))

    if "platform" in df.columns:
        platform_stats = df.groupby("platform").agg(
            order_count=("order_amount", "count"),
            total_revenue=("order_amount", "sum"),
            total_commission=("commission", "sum"),
            total_shipping=("shipping", "sum"),
            total_refund=("refund", "sum"),
        ).reset_index()

        platform_stats["gross_profit"] = (
            platform_stats["total_revenue"]
            - platform_stats["total_commission"]
            - platform_stats["total_shipping"]
        )
        platform_stats["gross_margin"] = platform_stats.apply(
            lambda r: r["gross_profit"] / r["total_revenue"] if r["total_revenue"] > 0 else 0,
            axis=1
        )
        platform_stats["net_profit"] = (
            platform_stats["total_revenue"]
            - platform_stats["total_commission"]
            - platform_stats["total_shipping"]
            - platform_stats["total_refund"]
        )

        number_cols2 = [3, 4, 5, 6, 7]
        for idx, p_row in platform_stats.iterrows():
            row += 1
            ws2.cell(row=row, column=1, value=str(p_row["platform"]))
            ws2.cell(row=row, column=2, value=int(p_row["order_count"]))
            ws2.cell(row=row, column=3, value=round(_safe_float(p_row["total_revenue"]), 2))
            ws2.cell(row=row, column=4, value=round(_safe_float(p_row["total_commission"]), 2))
            ws2.cell(row=row, column=5, value=round(_safe_float(p_row["total_shipping"]), 2))
            ws2.cell(row=row, column=6, value=round(_safe_float(p_row["total_refund"]), 2))
            ws2.cell(row=row, column=7, value=round(_safe_float(p_row["gross_profit"]), 2))
            ws2.cell(row=row, column=8, value=_safe_float(p_row["gross_margin"]))
            ws2.cell(row=row, column=8).number_format = _PERCENT_FORMAT
            ws2.cell(row=row, column=9, value=round(_safe_float(p_row["net_profit"]), 2))
            _apply_data_style(ws2, row, len(headers2), number_cols=number_cols2)

    _auto_width(ws2)

    # ==================================================================
    # Sheet3: SKU明细
    # ==================================================================
    ws3 = wb.create_sheet("SKU明细")

    ws3.merge_cells("A1:H1")
    ws3["A1"] = "SKU销售明细"
    ws3["A1"].font = _TITLE_FONT
    ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")

    row = 3
    headers3 = ["SKU", "平台", "订单数", "总收入", "总佣金", "总运费", "毛利润", "毛利率"]
    for col_idx, header in enumerate(headers3, 1):
        ws3.cell(row=row, column=col_idx, value=header)
    _apply_header_style(ws3, row, len(headers3))

    sku_col = "sku" if "sku" in df.columns else None
    plat_col = "platform" if "platform" in df.columns else None

    if sku_col:
        group_cols = [sku_col]
        if plat_col:
            group_cols.append(plat_col)

        sku_stats = df.groupby(group_cols).agg(
            order_count=("order_amount", "count"),
            total_revenue=("order_amount", "sum"),
            total_commission=("commission", "sum"),
            total_shipping=("shipping", "sum"),
        ).reset_index()

        sku_stats["gross_profit"] = (
            sku_stats["total_revenue"]
            - sku_stats["total_commission"]
            - sku_stats["total_shipping"]
        )
        sku_stats["gross_margin"] = sku_stats.apply(
            lambda r: r["gross_profit"] / r["total_revenue"] if r["total_revenue"] > 0 else 0,
            axis=1
        )

        # 按收入降序排列
        sku_stats = sku_stats.sort_values("total_revenue", ascending=False)

        number_cols3 = [4, 5, 6, 7]
        for idx, s_row in sku_stats.iterrows():
            row += 1
            ws3.cell(row=row, column=1, value=str(s_row[sku_col]))
            if plat_col:
                ws3.cell(row=row, column=2, value=str(s_row[plat_col]))
            else:
                ws3.cell(row=row, column=2, value="-")
            ws3.cell(row=row, column=3, value=int(s_row["order_count"]))
            ws3.cell(row=row, column=4, value=round(_safe_float(s_row["total_revenue"]), 2))
            ws3.cell(row=row, column=5, value=round(_safe_float(s_row["total_commission"]), 2))
            ws3.cell(row=row, column=6, value=round(_safe_float(s_row["total_shipping"]), 2))
            ws3.cell(row=row, column=7, value=round(_safe_float(s_row["gross_profit"]), 2))
            ws3.cell(row=row, column=8, value=_safe_float(s_row["gross_margin"]))
            ws3.cell(row=row, column=8).number_format = _PERCENT_FORMAT
            _apply_data_style(ws3, row, len(headers3), number_cols=number_cols3)

    _auto_width(ws3)

    # ==================================================================
    # Sheet4: 公式说明
    # ==================================================================
    ws4 = wb.create_sheet("公式说明")

    ws4.merge_cells("A1:D1")
    ws4["A1"] = "报表计算公式说明"
    ws4["A1"].font = _TITLE_FONT
    ws4["A1"].alignment = Alignment(horizontal="center", vertical="center")

    row = 3
    formulas = [
        ["指标", "计算公式", "说明", "注意事项"],
        ["总收入", "= SUM(订单金额)", "所有订单的销售金额总和", "按原始币种汇总"],
        ["总佣金", "= SUM(佣金/手续费)", "平台收取的佣金总和", "各平台佣金率不同"],
        ["总运费", "= SUM(运费)", "物流费用总和", "含FBA配送费"],
        ["总退款", "= SUM(退款金额)", "退款金额总和", "负值表示退款"],
        ["总成本", "= 总佣金 + 总运费 + 总退款", "所有成本费用总和", ""],
        ["毛利润", "= 总收入 - 总佣金 - 总运费", "扣除直接成本后的利润", "未扣除退款"],
        ["毛利率", "= 毛利润 / 总收入", "毛利润占收入的比例", "百分比格式"],
        ["净利润", "= 总收入 - 总成本", "扣除所有成本后的利润", "含退款"],
        ["净利率", "= 净利润 / 总收入", "净利润占收入的比例", "百分比格式"],
    ]

    for col_idx, header in enumerate(formulas[0], 1):
        ws4.cell(row=row, column=col_idx, value=header)
    _apply_header_style(ws4, row, 4)

    for formula_row in formulas[1:]:
        row += 1
        for col_idx, val in enumerate(formula_row, 1):
            ws4.cell(row=row, column=col_idx, value=val)
        _apply_data_style(ws4, row, 4)

    _auto_width(ws4)

    # 保存
    wb.save(output_path)
    logger.info(f"月度利润报表已生成: {output_path}")
    return output_path


# ===========================================================================
# 2. 对账差异报告
# ===========================================================================
def generate_reconciliation_report(
    matches: List[Dict[str, Any]],
    mismatches: List[Dict[str, Any]],
    output_path: str,
) -> str:
    """
    生成对账差异报告Excel。

    Args:
        matches: 匹配成功的记录列表，每项为字典
        mismatches: 匹配失败的记录列表，每项为字典
        output_path: 输出Excel文件路径

    Returns:
        str: 生成的文件路径
    """
    if not _OPENPYXL_AVAILABLE:
        raise ImportError("需要安装 openpyxl 库: pip install openpyxl")

    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else ".", exist_ok=True)

    wb = Workbook()

    # ---- Sheet1: 对账概要 ----
    ws1 = wb.active
    ws1.title = "对账概要"

    ws1.merge_cells("A1:D1")
    ws1["A1"] = "对账差异报告"
    ws1["A1"].font = _TITLE_FONT
    ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws1.merge_cells("A2:D2")
    ws1["A2"] = f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws1["A2"].font = Font(name="Microsoft YaHei", size=9, color="666666")
    ws1["A2"].alignment = Alignment(horizontal="center")

    row = 4
    summary_data = [
        ["指标", "数值"],
        ["匹配记录数", len(matches)],
        ["差异记录数", len(mismatches)],
        ["总记录数", len(matches) + len(mismatches)],
        ["匹配率", f"{len(matches) / max(len(matches) + len(mismatches), 1) * 100:.2f}%"],
    ]

    for col_idx, header in enumerate(summary_data[0], 1):
        ws1.cell(row=row, column=col_idx, value=header)
    _apply_header_style(ws1, row, 2)

    for data_row in summary_data[1:]:
        row += 1
        for col_idx, val in enumerate(data_row, 1):
            ws1.cell(row=row, column=col_idx, value=val)
        _apply_data_style(ws1, row, 2)

    _auto_width(ws1)

    # ---- Sheet2: 匹配记录 ----
    ws2 = wb.create_sheet("匹配记录")

    ws2.merge_cells("A1:F1")
    ws2["A1"] = f"匹配成功记录（共 {len(matches)} 条）"
    ws2["A1"].font = _SUBTITLE_FONT
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")

    if matches:
        row = 3
        match_keys = list(matches[0].keys())
        for col_idx, key in enumerate(match_keys, 1):
            ws2.cell(row=row, column=col_idx, value=str(key))
        _apply_header_style(ws2, row, len(match_keys))

        for match in matches:
            row += 1
            for col_idx, key in enumerate(match_keys, 1):
                val = match.get(key, "")
                ws2.cell(row=row, column=col_idx, value=str(val) if val is not None else "")
            _apply_data_style(ws2, row, len(match_keys))

    _auto_width(ws2)

    # ---- Sheet3: 差异记录 ----
    ws3 = wb.create_sheet("差异记录")

    ws3.merge_cells("A1:F1")
    ws3["A1"] = f"差异记录（共 {len(mismatches)} 条）"
    ws3["A1"].font = _SUBTITLE_FONT
    ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")

    # 差异记录用红色高亮
    _MISMATCH_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    if mismatches:
        row = 3
        mismatch_keys = list(mismatches[0].keys())
        for col_idx, key in enumerate(mismatch_keys, 1):
            ws3.cell(row=row, column=col_idx, value=str(key))
        _apply_header_style(ws3, row, len(mismatch_keys))

        for mismatch in mismatches:
            row += 1
            for col_idx, key in enumerate(mismatch_keys, 1):
                val = mismatch.get(key, "")
                cell = ws3.cell(row=row, column=col_idx, value=str(val) if val is not None else "")
                cell.fill = _MISMATCH_FILL
                cell.font = _DATA_FONT
                cell.border = _THIN_BORDER

    _auto_width(ws3)

    # 保存
    wb.save(output_path)
    logger.info(f"对账差异报告已生成: {output_path}")
    return output_path


# ===========================================================================
# 3. 图表生成（使用plotly）
# ===========================================================================
def create_revenue_chart(df: pd.DataFrame) -> Optional[str]:
    """
    生成收入趋势折线图（HTML格式）。

    Args:
        df: 包含订单数据的DataFrame

    Returns:
        str: HTML字符串，可直接嵌入网页；如果plotly不可用则返回None
    """
    if not _PLOTLY_AVAILABLE:
        logger.warning("plotly不可用，无法生成图表")
        return None

    df = df.copy()

    # 安全获取列（避免多列映射到同名冲突）
    def _find_col(candidates):
        for c in candidates:
            if c in df.columns:
                return c
        return None

    date_col = _find_col(["order_date", "date", "订单日期", "交易日期"])
    amt_col = _find_col(["order_amount", "amount", "订单金额", "销售额"])

    if not date_col or not amt_col:
        return None

    df["order_date"] = pd.to_datetime(df[date_col], errors="coerce")
    df["order_amount"] = pd.to_numeric(df[amt_col], errors="coerce").fillna(0)
    df = df.dropna(subset=["order_date"])

    # 按月汇总
    df["month"] = df["order_date"].dt.to_period("M")
    monthly = df.groupby("month")["order_amount"].sum().reset_index()
    monthly["month_str"] = monthly["month"].astype(str)

    fig = go.Figure()
    fig.add_trace(go.Scatter(
        x=monthly["month_str"],
        y=monthly["order_amount"],
        mode="lines+markers+text",
        line=dict(color="#4472C4", width=3),
        marker=dict(size=8),
        text=monthly["order_amount"].apply(lambda x: f"${x:,.2f}"),
        textposition="top center",
        name="月度收入",
    ))

    fig.update_layout(
        title="月度收入趋势",
        xaxis_title="月份",
        yaxis_title="收入金额",
        template="plotly_white",
        height=500,
        hovermode="x unified",
    )

    return fig.to_html(full_html=False, include_plotlyjs=False)


def create_cost_pie(df: pd.DataFrame) -> Optional[str]:
    """
    生成成本构成饼图（HTML格式）。

    Args:
        df: 包含订单数据的DataFrame

    Returns:
        str: HTML字符串；如果plotly不可用则返回None
    """
    if not _PLOTLY_AVAILABLE:
        logger.warning("plotly不可用，无法生成图表")
        return None

    df = df.copy()

    # 安全获取数值列（避免多列映射到同名冲突）
    def _find_col(candidates):
        for c in candidates:
            if c in df.columns:
                return c
        return None

    # 计算各项成本
    costs = {}
    com_col = _find_col(["commission", "fee", "佣金"])
    shp_col = _find_col(["shipping", "运费", "物流费"])
    ref_col = _find_col(["refund", "退款", "退款金额"])

    if com_col:
        costs["佣金/手续费"] = _safe_float(pd.to_numeric(df[com_col], errors="coerce").sum())
    if shp_col:
        costs["运费/物流费"] = _safe_float(pd.to_numeric(df[shp_col], errors="coerce").sum())
    if ref_col:
        costs["退款"] = _safe_float(pd.to_numeric(df[ref_col], errors="coerce").sum())

    if not costs:
        return None

    # 过滤掉为0的项
    costs = {k: v for k, v in costs.items() if v > 0}
    if not costs:
        return None

    labels = list(costs.keys())
    values = list(costs.values())

    fig = go.Figure(data=[go.Pie(
        labels=labels,
        values=values,
        hole=0.4,
        marker=dict(colors=_CHART_COLORS[:len(labels)]),
        textinfo="label+percent",
        textposition="outside",
    )])

    fig.update_layout(
        title="成本构成分析",
        template="plotly_white",
        height=500,
        showlegend=True,
    )

    return fig.to_html(full_html=False, include_plotlyjs=False)


def create_platform_comparison(df: pd.DataFrame) -> Optional[str]:
    """
    生成平台对比柱状图（HTML格式）。

    Args:
        df: 包含订单数据的DataFrame

    Returns:
        str: HTML字符串；如果plotly不可用则返回None
    """
    if not _PLOTLY_AVAILABLE:
        logger.warning("plotly不可用，无法生成图表")
        return None

    df = df.copy()

    # 安全获取列（避免多列映射到同名冲突）
    def _find_col(candidates):
        for c in candidates:
            if c in df.columns:
                return c
        return None

    plat_col = _find_col(["platform", "平台"])
    amt_col = _find_col(["order_amount", "amount", "订单金额", "销售额"])
    com_col = _find_col(["commission", "fee", "佣金"])

    if not plat_col or not amt_col:
        return None

    amounts = pd.to_numeric(df[amt_col], errors="coerce").fillna(0)
    commissions = pd.to_numeric(df[com_col], errors="coerce").fillna(0) if com_col else pd.Series(0, index=df.index)

    platform_stats = df.groupby(plat_col).agg(
        revenue=(amt_col, "sum"),
    ).reset_index()
    platform_stats["commission"] = df.groupby(plat_col).apply(
        lambda g: pd.to_numeric(g[com_col], errors="coerce").sum() if com_col else 0
    ).values
    platform_stats["orders"] = df.groupby(plat_col).size().values
    platform_stats = platform_stats.sort_values("revenue", ascending=True)

    fig = go.Figure()
    fig.add_trace(go.Bar(
        y=platform_stats["platform"],
        x=platform_stats["revenue"],
        orientation="h",
        name="总收入",
        marker=dict(color="#4472C4"),
        text=platform_stats["revenue"].apply(lambda x: f"${x:,.2f}"),
        textposition="outside",
    ))
    fig.add_trace(go.Bar(
        y=platform_stats["platform"],
        x=platform_stats["commission"],
        orientation="h",
        name="佣金",
        marker=dict(color="#ED7D31"),
        text=platform_stats["commission"].apply(lambda x: f"${x:,.2f}"),
        textposition="outside",
    ))

    fig.update_layout(
        title="各平台收入与佣金对比",
        xaxis_title="金额",
        yaxis_title="平台",
        barmode="group",
        template="plotly_white",
        height=max(400, len(platform_stats) * 50 + 100),
        showlegend=True,
    )

    return fig.to_html(full_html=False, include_plotlyjs=False)


def create_profit_gauge(profit_rate: float) -> Optional[str]:
    """
    生成利润率仪表盘（HTML格式）。

    Args:
        profit_rate: 利润率（小数形式，如 0.15 表示 15%）

    Returns:
        str: HTML字符串；如果plotly不可用则返回None
    """
    if not _PLOTLY_AVAILABLE:
        logger.warning("plotly不可用，无法生成图表")
        return None

    profit_rate = _safe_float(profit_rate)
    profit_pct = profit_rate * 100

    # 根据利润率确定颜色
    if profit_pct >= 20:
        color = "#70AD47"  # 绿色 - 优秀
    elif profit_pct >= 10:
        color = "#FFC000"  # 黄色 - 一般
    elif profit_pct >= 0:
        color = "#ED7D31"  # 橙色 - 较低
    else:
        color = "#FF0000"  # 红色 - 亏损

    fig = go.Figure(go.Indicator(
        mode="gauge+number",
        value=profit_pct,
        domain={"x": [0, 1], "y": [0, 1]},
        title={"text": "利润率"},
        number={"suffix": "%", "font": {"size": 40}},
        gauge={
            "axis": {
                "range": [-20, 50],
                "tickwidth": 1,
                "tickcolor": "#666666",
            },
            "bar": {"color": color, "thickness": 0.3},
            "bgcolor": "white",
            "borderwidth": 2,
            "bordercolor": "#CCCCCC",
            "steps": [
                {"range": [-20, 0], "color": "#FFE6E6"},
                {"range": [0, 10], "color": "#FFF2CC"},
                {"range": [10, 20], "color": "#E2EFDA"},
                {"range": [20, 50], "color": "#C6EFCE"},
            ],
            "threshold": {
                "line": {"color": "#333333", "width": 2},
                "thickness": 0.75,
                "value": profit_pct,
            },
        },
    ))

    fig.update_layout(
        title="利润率仪表盘",
        template="plotly_white",
        height=400,
        margin=dict(l=50, r=50, t=80, b=50),
    )

    return fig.to_html(full_html=False, include_plotlyjs=False)


# ===========================================================================
# 4. 一键生成完整分析报告
# ===========================================================================
def generate_full_report(
    df: pd.DataFrame,
    user_id: int,
    output_dir: str,
) -> Dict[str, str]:
    """
    一键生成完整分析报告，包含Excel数据 + HTML图表报告。

    Args:
        df: 包含订单数据的DataFrame
        user_id: 用户ID（用于文件命名）
        output_dir: 输出目录路径

    Returns:
        dict: 包含各输出文件路径的字典
            {
                "excel": "Excel报表路径",
                "html": "HTML图表报告路径",
                "charts": {
                    "revenue": "收入趋势图HTML片段",
                    "cost_pie": "成本饼图HTML片段",
                    "platform": "平台对比图HTML片段",
                    "profit_gauge": "利润率仪表盘HTML片段",
                }
            }
    """
    os.makedirs(output_dir, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    result = {
        "excel": "",
        "html": "",
        "charts": {},
    }

    # ---- 1. 生成Excel报表 ----
    excel_path = os.path.join(output_dir, f"profit_report_{user_id}_{timestamp}.xlsx")
    try:
        result["excel"] = generate_profit_report(df, excel_path)
    except Exception as e:
        logger.error(f"生成Excel报表失败: {e}")
        result["excel"] = ""

    # ---- 2. 生成图表HTML片段 ----
    try:
        revenue_chart = create_revenue_chart(df)
        result["charts"]["revenue"] = revenue_chart or ""
    except Exception as e:
        logger.error(f"生成收入趋势图失败: {e}")
        result["charts"]["revenue"] = ""

    try:
        cost_pie = create_cost_pie(df)
        result["charts"]["cost_pie"] = cost_pie or ""
    except Exception as e:
        logger.error(f"生成成本饼图失败: {e}")
        result["charts"]["cost_pie"] = ""

    try:
        platform_chart = create_platform_comparison(df)
        result["charts"]["platform"] = platform_chart or ""
    except Exception as e:
        logger.error(f"生成平台对比图失败: {e}")
        result["charts"]["platform"] = ""

    # 计算利润率
    try:
        df_temp = df.copy()
        # 安全获取数值列（避免多列映射冲突）
        _REV_COLS = ["order_amount", "amount", "订单金额", "销售额"]
        _COM_COLS = ["commission", "fee", "佣金"]
        _SHP_COLS = ["shipping", "运费", "物流费"]
        _REF_COLS = ["refund", "退款", "退款金额"]

        def _first_match(df, candidates):
            for c in candidates:
                if c in df.columns:
                    return pd.to_numeric(df[c], errors="coerce").fillna(0).sum()
            return 0.0

        total_revenue = _first_match(df_temp, _REV_COLS)
        total_cost = _first_match(df_temp, _COM_COLS) + _first_match(df_temp, _SHP_COLS) + _first_match(df_temp, _REF_COLS)
        net_profit = total_revenue - total_cost
        profit_rate = net_profit / total_revenue if total_revenue > 0 else 0
    except Exception:
        profit_rate = 0

    try:
        gauge = create_profit_gauge(profit_rate)
        result["charts"]["profit_gauge"] = gauge or ""
    except Exception as e:
        logger.error(f"生成利润率仪表盘失败: {e}")
        result["charts"]["profit_gauge"] = ""

    # ---- 3. 生成完整HTML报告 ----
    html_path = os.path.join(output_dir, f"analysis_report_{user_id}_{timestamp}.html")

    try:
        html_content = _build_html_report(
            df=df,
            user_id=user_id,
            charts=result["charts"],
            profit_rate=profit_rate,
            total_revenue=total_revenue,
            total_cost=total_cost,
            net_profit=net_profit,
        )

        with open(html_path, "w", encoding="utf-8") as f:
            f.write(html_content)

        result["html"] = html_path
        logger.info(f"HTML分析报告已生成: {html_path}")
    except Exception as e:
        logger.error(f"生成HTML报告失败: {e}")
        result["html"] = ""

    return result


def _build_html_report(
    df: pd.DataFrame,
    user_id: int,
    charts: Dict[str, str],
    profit_rate: float,
    total_revenue: float,
    total_cost: float,
    net_profit: float,
) -> str:
    """
    构建完整的HTML分析报告。

    Args:
        df: 原始数据DataFrame
        user_id: 用户ID
        charts: 图表HTML片段字典
        profit_rate: 利润率
        total_revenue: 总收入
        total_cost: 总成本
        net_profit: 净利润

    Returns:
        str: 完整的HTML字符串
    """
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # 计算统计信息
    order_count = len(df)

    # 平台统计
    platform_stats = {}
    if "platform" in df.columns:
        for plat, group in df.groupby("platform"):
            plat_revenue = 0
            for col in df.columns:
                if col.lower().strip() in ("order_amount", "amount", "订单金额", "销售额"):
                    plat_revenue = pd.to_numeric(group[col], errors="coerce").sum()
                    break
            platform_stats[str(plat)] = {
                "orders": len(group),
                "revenue": f"${plat_revenue:,.2f}",
                "share": f"{plat_revenue / total_revenue * 100:.1f}%" if total_revenue > 0 else "0%",
            }

    # 构建平台表格行
    platform_rows = ""
    for plat, stats in platform_stats.items():
        platform_rows += f"""
        <tr>
            <td>{plat}</td>
            <td>{stats['orders']}</td>
            <td>{stats['revenue']}</td>
            <td>{stats['share']}</td>
        </tr>"""

    # 利润率颜色
    if profit_rate >= 0.20:
        profit_color = "#28a745"
        profit_label = "优秀"
    elif profit_rate >= 0.10:
        profit_color = "#ffc107"
        profit_label = "良好"
    elif profit_rate >= 0:
        profit_color = "#fd7e14"
        profit_label = "偏低"
    else:
        profit_color = "#dc3545"
        profit_label = "亏损"

    html = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>财务分析报告 - 用户{user_id}</title>
    <script src="https://cdn.plot.ly/plotly-latest.min.js"></script>
    <style>
        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}
        body {{
            font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", "Microsoft YaHei", sans-serif;
            background: #f5f7fa;
            color: #333;
            line-height: 1.6;
        }}
        .container {{
            max-width: 1200px;
            margin: 0 auto;
            padding: 20px;
        }}
        .header {{
            background: linear-gradient(135deg, #1a73e8, #0d47a1);
            color: white;
            padding: 30px;
            border-radius: 12px;
            margin-bottom: 24px;
            box-shadow: 0 4px 12px rgba(0,0,0,0.1);
        }}
        .header h1 {{
            font-size: 28px;
            margin-bottom: 8px;
        }}
        .header .meta {{
            font-size: 14px;
            opacity: 0.85;
        }}
        .cards {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(220px, 1fr));
            gap: 16px;
            margin-bottom: 24px;
        }}
        .card {{
            background: white;
            border-radius: 10px;
            padding: 20px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.06);
            text-align: center;
        }}
        .card .label {{
            font-size: 13px;
            color: #666;
            margin-bottom: 8px;
        }}
        .card .value {{
            font-size: 24px;
            font-weight: 700;
        }}
        .card .sub {{
            font-size: 12px;
            color: #999;
            margin-top: 4px;
        }}
        .section {{
            background: white;
            border-radius: 10px;
            padding: 24px;
            margin-bottom: 24px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.06);
        }}
        .section h2 {{
            font-size: 20px;
            margin-bottom: 16px;
            padding-bottom: 12px;
            border-bottom: 2px solid #e8eaed;
            color: #1a73e8;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
        }}
        th, td {{
            padding: 12px 16px;
            text-align: left;
            border-bottom: 1px solid #e8eaed;
        }}
        th {{
            background: #f8f9fa;
            font-weight: 600;
            color: #555;
        }}
        tr:hover {{
            background: #f0f7ff;
        }}
        .chart-container {{
            margin: 16px 0;
        }}
        .footer {{
            text-align: center;
            padding: 20px;
            color: #999;
            font-size: 12px;
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>财务分析报告</h1>
            <div class="meta">
                用户ID: {user_id} | 生成时间: {now_str} | 数据记录: {order_count} 条
            </div>
        </div>

        <div class="cards">
            <div class="card">
                <div class="label">总收入</div>
                <div class="value" style="color: #1a73e8;">${total_revenue:,.2f}</div>
                <div class="sub">全部订单金额</div>
            </div>
            <div class="card">
                <div class="label">总成本</div>
                <div class="value" style="color: #dc3545;">${total_cost:,.2f}</div>
                <div class="sub">佣金+运费+退款</div>
            </div>
            <div class="card">
                <div class="label">净利润</div>
                <div class="value" style="color: {profit_color};">${net_profit:,.2f}</div>
                <div class="sub">收入-成本</div>
            </div>
            <div class="card">
                <div class="label">净利率</div>
                <div class="value" style="color: {profit_color};">{profit_rate * 100:.2f}%</div>
                <div class="sub">{profit_label}</div>
            </div>
        </div>

        <div class="section">
            <h2>利润率仪表盘</h2>
            <div class="chart-container">
                {charts.get("profit_gauge", "<p>图表生成失败</p>")}
            </div>
        </div>

        <div class="section">
            <h2>收入趋势</h2>
            <div class="chart-container">
                {charts.get("revenue", "<p>图表生成失败</p>")}
            </div>
        </div>

        <div class="section">
            <h2>成本构成</h2>
            <div class="chart-container">
                {charts.get("cost_pie", "<p>图表生成失败</p>")}
            </div>
        </div>

        <div class="section">
            <h2>平台对比</h2>
            <div class="chart-container">
                {charts.get("platform", "<p>图表生成失败</p>")}
            </div>
        </div>

        <div class="section">
            <h2>平台明细</h2>
            <table>
                <thead>
                    <tr>
                        <th>平台</th>
                        <th>订单数</th>
                        <th>收入</th>
                        <th>占比</th>
                    </tr>
                </thead>
                <tbody>
                    {platform_rows if platform_rows else "<tr><td colspan='4'>暂无平台数据</td></tr>"}
                </tbody>
            </table>
        </div>

        <div class="footer">
            <p>跨境财务助手 - 自动生成报告 | 数据仅供参考，请以实际账单为准</p>
        </div>
    </div>
</body>
</html>"""

    return html


# ===========================================================================
# 模块测试入口
# ===========================================================================
if __name__ == "__main__":
    print("=" * 60)
    print("  报表生成器模块测试")
    print("=" * 60)

    # 创建测试数据
    import numpy as np

    np.random.seed(42)
    dates = pd.date_range("2024-01-01", periods=90, freq="D")
    platforms = ["Amazon", "eBay", "Shopee", "Shopify"]
    skus = ["SKU-A001", "SKU-B002", "SKU-C003", "SKU-D004", "SKU-E005"]

    test_data = {
        "order_date": np.random.choice(dates, 200),
        "platform": np.random.choice(platforms, 200),
        "sku": np.random.choice(skus, 200),
        "order_amount": np.random.uniform(10, 500, 200).round(2),
        "commission": np.random.uniform(1, 50, 200).round(2),
        "shipping": np.random.uniform(2, 20, 200).round(2),
        "refund": np.random.choice([0, 0, 0, 0, np.random.uniform(5, 30)], 200).round(2),
        "currency": "USD",
    }
    test_df = pd.DataFrame(test_data)

    # 测试Excel报表生成
    print("\n[1] 测试月度利润报表生成...")
    output_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "test_reports")
    excel_path = os.path.join(output_dir, "test_profit_report.xlsx")
    try:
        result = generate_profit_report(test_df, excel_path)
        print(f"  Excel报表已生成: {result}")
    except Exception as e:
        print(f"  生成失败: {e}")

    # 测试对账报告生成
    print("\n[2] 测试对账差异报告生成...")
    matches = [
        {"order_id": "ORD001", "amount": 100.00, "status": "matched"},
        {"order_id": "ORD002", "amount": 200.00, "status": "matched"},
    ]
    mismatches = [
        {"order_id": "ORD003", "expected": 150.00, "actual": 145.00, "diff": -5.00},
    ]
    recon_path = os.path.join(output_dir, "test_reconciliation.xlsx")
    try:
        result = generate_reconciliation_report(matches, mismatches, recon_path)
        print(f"  对账报告已生成: {result}")
    except Exception as e:
        print(f"  生成失败: {e}")

    # 测试图表生成
    print("\n[3] 测试图表生成...")
    chart_funcs = [
        ("收入趋势图", create_revenue_chart),
        ("成本饼图", create_cost_pie),
        ("平台对比图", create_platform_comparison),
        ("利润率仪表盘", lambda df: create_profit_gauge(0.15)),
    ]
    for name, func in chart_funcs:
        try:
            result = func(test_df)
            status = "成功" if result else "跳过(plotly不可用)"
            print(f"  {name}: {status}")
        except Exception as e:
            print(f"  {name}: 失败 - {e}")

    # 测试完整报告
    print("\n[4] 测试完整报告生成...")
    try:
        result = generate_full_report(test_df, 1, output_dir)
        print(f"  Excel: {result.get('excel', '未生成')}")
        print(f"  HTML: {result.get('html', '未生成')}")
    except Exception as e:
        print(f"  生成失败: {e}")

    print("\n" + "=" * 60)
    print("  测试完成")
    print("=" * 60)
